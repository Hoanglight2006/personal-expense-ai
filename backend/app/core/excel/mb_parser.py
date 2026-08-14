import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl

from app.core.excel.base_parser import ExcelParser
from app.models.enums import CategoryType, PaymentMethod
from app.schemas.transaction import BulkTransactionRow


class MBStatementAdapter(ExcelParser):
    """Parser specifically for MBBank account statements.
    
    Rules based on user request:
    - Identify by sheet name and header columns
    - Header VN at row 19, EN at row 20 (or dynamic detection)
    - Data starts at 21 (or after headers)
    - Column mapping:
      A: Ngày giao dịch
      B: Ngày hạch toán
      C: Số bút toán
      D: Phát sinh nợ (Chi)
      E: Phát sinh có (Thu)
      F: Số dư lũy kế (Ignore)
      G: Nội dung
      H: Đơn vị thụ hưởng
      I: Tài khoản
      J: Ngân hàng đối tác
    - Default payment method: BANK_TRANSFER
    """

    def can_parse(self, file_content: bytes) -> bool:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            # MB statements typically have 'Sao ke tai khoan' or similar sheet names
            if "Sao ke tai khoan" in wb.sheetnames:
                return True
            # Fallback: check first sheet for MB specific headers
            sheet = wb.active
            for row in sheet.iter_rows(min_row=1, max_row=30, values_only=True):
                if row and len(row) >= 7:
                    row_str = " ".join([str(c).lower() for c in row if c])
                    if "số bút toán" in row_str or "transaction no" in row_str:
                        return True
            return False
        except Exception:
            return False

    def parse(self, file_content: bytes) -> list[BulkTransactionRow]:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = wb["Sao ke tai khoan"] if "Sao ke tai khoan" in wb.sheetnames else wb.active

        rows = []
        data_started = False
        
        for excel_row in sheet.iter_rows(values_only=True):
            if not excel_row or not any(excel_row):
                continue
            
            row_str = " ".join([str(c).lower() for c in excel_row if c])
            
            # Detect headers to skip
            if "số bút toán" in row_str or "transaction no" in row_str:
                data_started = True
                continue
            if "tổng phát sinh trong kỳ" in row_str or "số dư cuối kỳ" in row_str or "total" in row_str:
                break # Footer reached
                
            if not data_started:
                continue

            # Ensure row has enough columns
            if len(excel_row) < 7:
                continue

            # Parse columns
            date_val = excel_row[0]
            ref_no = str(excel_row[2]) if excel_row[2] else ""
            debit = excel_row[3]
            credit = excel_row[4]
            desc = str(excel_row[6]) if excel_row[6] else ""
            
            partner = str(excel_row[7]) if len(excel_row) > 7 and excel_row[7] else ""
            account = str(excel_row[8]) if len(excel_row) > 8 and excel_row[8] else ""
            bank = str(excel_row[9]) if len(excel_row) > 9 and excel_row[9] else ""
            
            # Construct full description
            full_desc = desc
            if partner or account or bank:
                extras = [p for p in [partner, account, bank] if p and p != "None"]
                if extras:
                    full_desc += f" | Đối tác: {', '.join(extras)}"
            full_desc = full_desc[:255]

            # Parse amount and type
            debit_val = self._parse_amount(debit)
            credit_val = self._parse_amount(credit)
            
            if debit_val > 0 and credit_val > 0:
                # Both have values -> Error/Skip
                continue
            if debit_val == 0 and credit_val == 0:
                continue

            txn_type = CategoryType.EXPENSE if debit_val > 0 else CategoryType.INCOME
            amount = debit_val if debit_val > 0 else credit_val
            
            # Parse date
            txn_date = self._parse_date(date_val)
            if not txn_date:
                continue

            # Ref no will be part of the fingerprint later, we encode it in description for now or just trust the detector
            if ref_no and ref_no != "None":
                full_desc = f"[{ref_no}] {full_desc}"[:255]

            # Note: category_id is set to 0 as a placeholder, it will be mapped later by category_suggestion
            rows.append(BulkTransactionRow(
                amount=amount,
                type=txn_type,
                category_id=0,
                transaction_date=txn_date,
                description=full_desc,
                payment_method=PaymentMethod.BANK_TRANSFER,
            ))

        return rows

    def _parse_amount(self, val) -> Decimal:
        if not val:
            return Decimal(0)
        try:
            if isinstance(val, (int, float)):
                return Decimal(str(val))
            clean_val = str(val).replace(",", "")
            return Decimal(clean_val)
        except InvalidOperation:
            return Decimal(0)

    def _parse_date(self, val) -> datetime.date | None:
        if isinstance(val, datetime):
            return val.date()
        if not val:
            return None
        # Try parse string like DD/MM/YYYY
        try:
            return datetime.strptime(str(val).strip()[:10], "%d/%m/%Y").date()
        except ValueError:
            try:
                return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
            except ValueError:
                return None
